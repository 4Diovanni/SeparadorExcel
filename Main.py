# separador de casos por área -----------------------------------------------------

# -*- coding: utf-8 -*-
"""
Separador de casos por área
Automação desenvolvida por: Giovanni Moreira
Assistente financeiro da empresa Allcare Gestora de Saúde
Data: 09/05/2025
version: 1.6.2
Descrição: Este código tem como objetivo automatizar a separação de arquivos por área, garantindo que cada arquivo seja movido para a pasta correta, com o mesmo nome da respectiva área.
"""

import openpyxl
from openpyxl import Workbook
import os
from tkinter import Tk, filedialog
from Separador import *

# Função para selecionar uma pasta
def select_folder():
    try:
        root = Tk()
        root.withdraw()
        folder_selected = filedialog.askdirectory(title="Selecione a Pasta")
        if not folder_selected:
            raise Exception("Nenhuma pasta selecionada.")
        return folder_selected
    except Exception as e:
        print(f"Erro ao selecionar a pasta: {e}")
        return None

# Função para criar um novo arquivo
def criar_arquivo(caminho, novo):
    try:
        wb = Workbook()
        wb.save(os.path.join(caminho, novo))
    except Exception as e:
        print(f"Erro ao criar o arquivo {novo}: {e}")

# Função para preparar a classificação
def preparar(ws, col_base):
    try:
        ws.auto_filter.ref = ws.dimensions
        ws.auto_filter.add_sort_condition(f"{col_base}1:{col_base}{ws.max_row}")
    except Exception as e:
        print(f"Erro ao preparar a classificação: {e}")

# Função para separar os arquivos
def separar(ws, col_base, pasta):
    try:
        grupos = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            valor = row[ws[col_base + '1'].column - 1]
            if valor not in grupos:
                grupos[valor] = []
            grupos[valor].append(row)

        for valor, linhas in grupos.items():
            novo_wb = Workbook()
            novo_ws = novo_wb.active

            # Copia o cabeçalho
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                novo_ws.append(row)# type: ignore

            # Copia as linhas de dados
            for linha in linhas:
                novo_ws.append(linha)# type: ignore

            # Salva o novo arquivo
            novo_wb.save(os.path.join(pasta, f"{valor}.xlsx")) # ATENÇÃO!!! o arquivo está no formato .xlsx!!! pode ser alterado.

        print("Planilhas geradas com sucesso!")
    except Exception as e:
        print(f"Erro ao separar os arquivos: {e}")

# Função para consolidar os arquivos separados em um único arquivo
def consolidar_arquivos(pasta):
    try:
        consolidado_wb = Workbook()
        consolidado_ws = consolidado_wb.active

        # Adiciona cabeçalho ao consolidado
        primeiro_arquivo = True

        for filename in os.listdir(pasta):
            if filename.endswith(".xlsx"):
                filepath = os.path.join(pasta, filename)
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active

                if primeiro_arquivo:
                    # Copia o cabeçalho do primeiro arquivo
                    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):# type: ignore
                        consolidado_ws.append(row)# type: ignore
                    primeiro_arquivo = False

                # Copia as linhas de dados dos arquivos separados
                for row in ws.iter_rows(min_row=2, values_only=True):# type: ignore
                    consolidado_ws.append(row) # type: ignore

        # Salva o arquivo consolidado
        consolidado_wb.save(os.path.join(pasta, "consolidado.xlsx"))
        print("Arquivo consolidado gerado com sucesso!")
    except Exception as e:
        print(f"Erro ao consolidar os arquivos: {e}")

# Exemplo de uso
try:
    pasta = select_folder()
    if pasta:
        wb = openpyxl.load_workbook(r'Caminho/do/arquivo/para/separar.xlsx')
        ws = wb.active
        preparar(ws, 'AH') # selecione a coluna por aqui!
        separar(ws, 'AH', pasta) # selecione a coluna por aqui!
        
        # Consolidar os arquivos separados em um único arquivo
        consolidar_arquivos(pasta)
        separador_pasta()
        print("Separação e consolidação concluídas com sucesso!")
        
except Exception as e:
    print(f"Erro ao executar o script: {e}")